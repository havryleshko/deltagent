from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def write_workbook(path: Path, company: str, accounts: list[tuple[str, int, int]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Variance"

    ws["A1"] = company
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Budget Variance"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = "1 November 2024"
    ws["A3"].font = Font(size=11)

    ws["A5"] = "Account"
    ws["B5"] = "Actual"
    ws["C5"] = "Budget"
    for cell in ws["A5:C5"][0]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right")

    for idx, (account, budget, actual) in enumerate(accounts, start=6):
        ws.cell(row=idx, column=1, value=account)
        ws.cell(row=idx, column=2, value=actual)
        ws.cell(row=idx, column=3, value=budget)
        ws.cell(row=idx, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=idx, column=3).alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    wb.save(path)


def supported_line(
    line_item: str,
    budget: int,
    actual: int,
    expected_driver: str,
    driver_keywords: list[str],
    allowed_source_families: list[str],
    evidence_ids: list[str],
    mitigation: str = "",
    forward_risk: str = "",
) -> dict:
    variance = actual - budget
    variance_pct = round((variance / budget) * 100, 4) if budget else 0.0
    payload = {
        "line_item": line_item,
        "budget_usd": budget,
        "actual_usd": actual,
        "variance_usd": variance,
        "variance_pct": variance_pct,
        "significant": True,
        "supported": True,
        "expected_driver": expected_driver,
        "expected_driver_keywords": driver_keywords,
        "allowed_source_families": allowed_source_families,
        "expected_evidence_ids": evidence_ids,
    }
    if mitigation:
        payload["mitigation_or_action"] = mitigation
    if forward_risk:
        payload["forward_risk"] = forward_risk
    return payload


def unsupported_line(line_item: str, budget: int, actual: int) -> dict:
    variance = actual - budget
    variance_pct = round((variance / budget) * 100, 4) if budget else 0.0
    return {
        "line_item": line_item,
        "budget_usd": budget,
        "actual_usd": actual,
        "variance_usd": variance,
        "variance_pct": variance_pct,
        "significant": True,
        "supported": False,
    }


def build_specs() -> list[dict]:
    return [
        {
            "slug": "bluebird_biotech_services_november_2024",
            "company": "Bluebird Biotech Services",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Salaries", "Professional Fees", "Cold Chain Logistics"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 420000, 468500),
                ("Grant Revenue", 64000, 60300),
                ("Salaries", 182000, 201500),
                ("Professional Fees", 26000, 41800),
                ("Cold Chain Logistics", 54000, 64400),
                ("Lab Consumables", 88000, 90100),
                ("Software & Subscriptions", 19000, 19850),
                ("Clinical Trial Support", 72000, 68900),
                ("Insurance", 13000, 12850),
                ("Travel & Training", 11000, 12100),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM renewal board shows two enterprise expansions signed in the final week of November worth +32000 versus plan.",
                        "narrow": "Opportunity audit trail shows one December close pulled into November after legal redlines cleared early.",
                    }
                },
                "search_gmail": {
                    "Salaries": {
                        "broad": "CFO approval email confirms six QA contractors converted to payroll effective 11 November.",
                    },
                    "Professional Fees": {
                        "broad": "Invoice chain confirms one-off validation counsel and external quality audit support posted in November.",
                        "narrow": "AP note ties 11800 of spend to remediation work for a delayed customer validation package.",
                    },
                },
                "search_slack": {
                    "Salaries": {
                        "broad": "People Ops update notes overlap payroll for contractor conversion and one backfill start in the same month.",
                    },
                    "Revenue": {
                        "narrow": "Commercial channel confirms expansion paperwork closed before month-end after redlines cleared.",
                    },
                },
                "search_calendar": {
                    "Professional Fees": {
                        "broad": "Validation readiness review with external counsel and quality advisor scheduled during week three of November.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    420000,
                    468500,
                    "enterprise expansions and one pulled-forward deal lifted November revenue",
                    ["enterprise expansions", "pulled into November", "redlines", "month-end"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "crm-revenue-narrow", "slack-revenue-narrow"],
                    mitigation="check whether December pipeline is lighter after the pull-forward",
                    forward_risk="December pipeline may soften after the accelerated close",
                ),
                supported_line(
                    "Salaries",
                    182000,
                    201500,
                    "contractor conversions and backfill overlap increased payroll",
                    ["converted to payroll", "overlap payroll", "backfill start"],
                    ["gmail", "slack"],
                    ["gmail-salaries-broad", "slack-salaries-broad"],
                ),
                supported_line(
                    "Professional Fees",
                    26000,
                    41800,
                    "one-off validation counsel and quality audit support drove the overspend",
                    ["validation counsel", "quality audit", "remediation work"],
                    ["gmail", "calendar"],
                    ["gmail-professional_fees-broad", "gmail-professional_fees-narrow", "calendar-professional_fees-broad"],
                ),
                unsupported_line("Cold Chain Logistics", 54000, 64400),
            ],
        },
        {
            "slug": "cedar_facilities_management_november_2024",
            "company": "Cedar Facilities Management",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Repairs & Maintenance", "Security Services", "Fuel & Travel"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 510000, 548200),
                ("Salaries", 206000, 221400),
                ("Repairs & Maintenance", 118000, 142900),
                ("Security Services", 63000, 74400),
                ("Fuel & Travel", 21000, 24750),
                ("Professional Fees", 22000, 23400),
                ("Cleaning Supplies", 38000, 37400),
                ("Software & Subscriptions", 17000, 16300),
                ("Insurance", 16000, 15850),
                ("Recruiting", 9000, 9800),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM pipeline shows two municipal contract expansions activated in November adding +27000 to monthly billing.",
                    }
                },
                "search_gmail": {
                    "Repairs & Maintenance": {
                        "broad": "Vendor invoice batch covers emergency HVAC callouts across three managed sites after cold-weather failures.",
                        "narrow": "Operations director approved 9600 of after-hours repair labor for priority sites.",
                    }
                },
                "search_slack": {
                    "Security Services": {
                        "broad": "Regional ops channel notes extra overnight guard coverage added for two vacant transition sites.",
                    },
                    "Revenue": {
                        "narrow": "Account team confirms municipal add-on scope went live before month-end billing cut-off.",
                    }
                },
                "search_calendar": {
                    "Security Services": {
                        "broad": "Transition-site risk review scheduled weekly through late November with vendor security lead.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    510000,
                    548200,
                    "municipal contract expansions activated before the billing cut-off",
                    ["contract expansions", "live before month-end", "billing cut-off"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "slack-revenue-narrow"],
                ),
                supported_line(
                    "Repairs & Maintenance",
                    118000,
                    142900,
                    "emergency hvac callouts and after-hours repair labor drove the overspend",
                    ["emergency hvac", "after-hours repair labor", "cold-weather failures"],
                    ["gmail"],
                    ["gmail-repairs_&_maintenance-broad", "gmail-repairs_&_maintenance-narrow"],
                ),
                supported_line(
                    "Security Services",
                    63000,
                    74400,
                    "extra overnight guard coverage for transition sites increased spend",
                    ["overnight guard coverage", "transition sites", "risk review"],
                    ["slack", "calendar"],
                    ["slack-security_services-broad", "calendar-security_services-broad"],
                ),
                unsupported_line("Fuel & Travel", 21000, 24750),
            ],
        },
        {
            "slug": "delta_logistics_network_november_2024",
            "company": "Delta Logistics Network",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Warehouse Labor", "Professional Fees", "Fleet Repairs"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 690000, 748500),
                ("Warehouse Labor", 155000, 176400),
                ("Professional Fees", 14000, 26300),
                ("Fleet Repairs", 82000, 96350),
                ("Fuel", 108000, 117400),
                ("Freight Claims", 9000, 8400),
                ("Software & Subscriptions", 21000, 20950),
                ("Insurance", 18000, 18700),
                ("Recruiting", 7000, 9450),
                ("Travel & Entertainment", 8500, 8800),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM account notes confirm a consumer electronics surge order and one regional contract expansion added +41000 to November revenue.",
                    }
                },
                "search_gmail": {
                    "Professional Fees": {
                        "broad": "Outside customs counsel invoice posted in November for urgent port compliance dispute support.",
                    },
                    "Warehouse Labor": {
                        "broad": "Staffing vendor summary shows overtime coverage and temporary dock labor added during two peak weeks.",
                    }
                },
                "search_slack": {
                    "Warehouse Labor": {
                        "narrow": "Ops channel says inbound surge required Saturday shifts across both hubs for two weekends.",
                    },
                    "Revenue": {
                        "narrow": "Sales thread confirms surge order shipped before Thanksgiving and billed in November.",
                    }
                },
                "search_calendar": {
                    "Professional Fees": {
                        "broad": "Port compliance dispute review with customs counsel held mid-November.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    690000,
                    748500,
                    "a surge order and regional expansion lifted november revenue",
                    ["surge order", "regional expansion", "billed in november"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "slack-revenue-narrow"],
                ),
                supported_line(
                    "Warehouse Labor",
                    155000,
                    176400,
                    "peak-week overtime and temporary dock labor drove warehouse payroll over budget",
                    ["overtime coverage", "temporary dock labor", "saturday shifts", "peak weeks"],
                    ["gmail", "slack"],
                    ["gmail-warehouse_labor-broad", "slack-warehouse_labor-narrow"],
                ),
                supported_line(
                    "Professional Fees",
                    14000,
                    26300,
                    "urgent port compliance dispute support from customs counsel caused the spike",
                    ["customs counsel", "port compliance dispute"],
                    ["gmail", "calendar"],
                    ["gmail-professional_fees-broad", "calendar-professional_fees-broad"],
                ),
                unsupported_line("Fleet Repairs", 82000, 96350),
            ],
        },
        {
            "slug": "ember_saas_security_november_2024",
            "company": "Ember SaaS Security",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Cloud Hosting", "Professional Fees", "Sales Commissions"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 560000, 603400),
                ("Cloud Hosting", 94000, 112500),
                ("Professional Fees", 18000, 31200),
                ("Sales Commissions", 52000, 61150),
                ("Salaries", 248000, 256500),
                ("Customer Success Tools", 21000, 22800),
                ("Software & Subscriptions", 17000, 18200),
                ("Travel & Entertainment", 12000, 11850),
                ("Insurance", 22000, 21900),
                ("Recruiting", 14000, 14950),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM shows one multi-year enterprise win and one mid-market upsell closing in the final week of November.",
                    }
                },
                "search_gmail": {
                    "Cloud Hosting": {
                        "broad": "AWS invoice summary shows temporary compute spike from an accelerated customer data migration and staging overlap.",
                        "narrow": "FinOps note says duplicate staging environment ran for nine days during migration cutover.",
                    },
                    "Professional Fees": {
                        "broad": "External incident-response counsel and privacy review invoices were approved after a customer escalation.",
                    }
                },
                "search_slack": {
                    "Sales Commissions": {
                        "broad": "RevOps confirms commission accrual increased with the enterprise win and mid-market expansion booked in November.",
                    },
                    "Revenue": {
                        "narrow": "Deal desk thread confirms enterprise win was signed before month-end and commissions were accrued in the same month.",
                    }
                },
                "search_calendar": {
                    "Professional Fees": {
                        "broad": "Privacy escalation review with outside counsel held on 19 November.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    560000,
                    603400,
                    "an enterprise win and a mid-market upsell closed before month-end",
                    ["enterprise win", "upsell", "signed before month-end"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "slack-revenue-narrow"],
                ),
                supported_line(
                    "Cloud Hosting",
                    94000,
                    112500,
                    "accelerated migration compute and duplicate staging environment drove cloud overage",
                    ["accelerated customer data migration", "duplicate staging environment", "migration cutover"],
                    ["gmail"],
                    ["gmail-cloud_hosting-broad", "gmail-cloud_hosting-narrow"],
                ),
                supported_line(
                    "Professional Fees",
                    18000,
                    31200,
                    "external incident-response counsel and privacy review spend caused the spike",
                    ["incident-response counsel", "privacy review", "customer escalation"],
                    ["gmail", "calendar"],
                    ["gmail-professional_fees-broad", "calendar-professional_fees-broad"],
                ),
                supported_line(
                    "Sales Commissions",
                    52000,
                    61150,
                    "commission accrual rose with the enterprise win and upsell bookings",
                    ["commission accrual", "enterprise win", "upsell bookings"],
                    ["slack"],
                    ["slack-sales_commissions-broad"],
                ),
            ],
        },
        {
            "slug": "forge_industrial_components_november_2024",
            "company": "Forge Industrial Components",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Raw Materials", "Repairs & Maintenance", "Security — Plant"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 780000, 752400),
                ("Raw Materials", 244000, 274800),
                ("Repairs & Maintenance", 86000, 101900),
                ("Security — Plant", 39000, 45400),
                ("Salaries", 198000, 203200),
                ("Professional Fees", 15000, 14900),
                ("Utilities", 71000, 76400),
                ("Freight", 44000, 47000),
                ("Insurance", 16000, 15950),
                ("Travel & Training", 9500, 9800),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM notes show one customer delayed acceptance into December, reducing November shipment recognition by 24000.",
                    }
                },
                "search_gmail": {
                    "Raw Materials": {
                        "broad": "Steel supplier notice confirms a spot-buy at elevated pricing after mill allocation tightened mid-month.",
                    },
                    "Repairs & Maintenance": {
                        "broad": "Maintenance invoice packet shows emergency press repairs and weekend technician callout charges.",
                    }
                },
                "search_slack": {
                    "Security — Plant": {
                        "broad": "Plant ops channel notes temporary gate coverage added after a weekend trespass incident.",
                    },
                    "Revenue": {
                        "narrow": "Sales ops confirms customer acceptance paperwork slipped to 2 December, deferring shipment revenue.",
                    }
                },
                "search_calendar": {
                    "Repairs & Maintenance": {
                        "broad": "Emergency vendor service visit logged for primary press line during week two of November.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    780000,
                    752400,
                    "customer acceptance slipped into december and deferred shipment revenue",
                    ["delayed acceptance", "slipped to december", "deferring shipment revenue"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "slack-revenue-narrow"],
                    mitigation="confirm whether December recovered the deferred shipment",
                ),
                supported_line(
                    "Raw Materials",
                    244000,
                    274800,
                    "spot-buy steel at elevated pricing drove raw material inflation",
                    ["spot-buy", "elevated pricing", "allocation tightened"],
                    ["gmail"],
                    ["gmail-raw_materials-broad"],
                ),
                supported_line(
                    "Repairs & Maintenance",
                    86000,
                    101900,
                    "emergency press repairs and weekend technician callouts caused the overspend",
                    ["emergency press repairs", "weekend technician callout", "service visit"],
                    ["gmail", "calendar"],
                    ["gmail-repairs_&_maintenance-broad", "calendar-repairs_&_maintenance-broad"],
                ),
                supported_line(
                    "Security — Plant",
                    39000,
                    45400,
                    "temporary gate coverage after a trespass incident increased security costs",
                    ["temporary gate coverage", "trespass incident"],
                    ["slack"],
                    ["slack-security_—_plant-broad"],
                ),
            ],
        },
        {
            "slug": "glenhaven_hospitality_group_november_2024",
            "company": "Glenhaven Hospitality Group",
            "summary_expectations": {
                "must_surface_lines": ["Room Revenue", "Food & Beverage Revenue", "Kitchen Payroll", "Property Insurance"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Room Revenue", 390000, 421500),
                ("Food & Beverage Revenue", 74000, 85100),
                ("Kitchen Payroll", 48000, 55750),
                ("Property Insurance", 31000, 35300),
                ("Housekeeping Payroll", 69000, 71500),
                ("Marketing Programs", 22000, 26900),
                ("Utilities", 41000, 42600),
                ("Linen & Laundry", 16000, 17100),
                ("Professional Fees", 9000, 9700),
                ("Travel & Entertainment", 6000, 5750),
            ],
            "mock_context": {
                "search_crm": {
                    "Room Revenue": {
                        "broad": "Group booking report shows one wedding block and one corporate retreat added 220 room nights above plan in November.",
                    }
                },
                "search_gmail": {
                    "Property Insurance": {
                        "broad": "Broker renewal summary confirms higher property premium after updated coastal catastrophe modeling.",
                    }
                },
                "search_slack": {
                    "Food & Beverage Revenue": {
                        "broad": "F&B channel says banquet package upgrades and add-on beverage minimums lifted event revenue during two weekends.",
                    },
                    "Kitchen Payroll": {
                        "broad": "Chef scheduling update notes overtime and agency prep cooks added for peak banquet weekends.",
                    }
                },
                "search_calendar": {
                    "Food & Beverage Revenue": {
                        "narrow": "Events calendar shows two expanded wedding receptions and one corporate banquet during the final two weekends.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Room Revenue",
                    390000,
                    421500,
                    "a wedding block and corporate retreat lifted room nights above plan",
                    ["wedding block", "corporate retreat", "room nights"],
                    ["crm"],
                    ["crm-room_revenue-broad"],
                ),
                supported_line(
                    "Food & Beverage Revenue",
                    74000,
                    85100,
                    "banquet package upgrades and beverage minimums lifted event revenue",
                    ["banquet package upgrades", "beverage minimums", "expanded receptions"],
                    ["slack", "calendar"],
                    ["slack-food_&_beverage_revenue-broad", "calendar-food_&_beverage_revenue-narrow"],
                ),
                supported_line(
                    "Kitchen Payroll",
                    48000,
                    55750,
                    "overtime and agency prep cooks for banquet weekends increased payroll",
                    ["overtime", "agency prep cooks", "banquet weekends"],
                    ["slack"],
                    ["slack-kitchen_payroll-broad"],
                ),
                supported_line(
                    "Property Insurance",
                    31000,
                    35300,
                    "broker renewal increased premiums after revised catastrophe modeling",
                    ["broker renewal", "coastal catastrophe modeling", "higher property premium"],
                    ["gmail"],
                    ["gmail-property_insurance-broad"],
                ),
            ],
        },
        {
            "slug": "harborline_health_clinics_november_2024",
            "company": "Harborline Health Clinics",
            "summary_expectations": {
                "must_surface_lines": ["Patient Service Revenue", "Contract Labor — Clinical", "Medical Supplies", "Medical Compliance & Audit"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Patient Service Revenue", 620000, 671500),
                ("Contract Labor — Clinical", 134000, 156900),
                ("Medical Supplies", 97000, 109300),
                ("Medical Compliance & Audit", 18000, 24600),
                ("Salaries", 212000, 228500),
                ("Professional Fees", 16000, 18500),
                ("Insurance", 19000, 20100),
                ("Travel & Training", 8500, 9100),
                ("Software & Subscriptions", 24000, 24350),
                ("Office & Facilities", 17000, 17550),
            ],
            "mock_context": {
                "search_crm": {
                    "Patient Service Revenue": {
                        "broad": "Referral dashboard shows ortho procedure volume above plan after two surgeon schedules reopened in November.",
                    }
                },
                "search_gmail": {
                    "Medical Compliance & Audit": {
                        "broad": "External compliance review invoice approved following an unplanned payer documentation audit.",
                    },
                    "Contract Labor — Clinical": {
                        "broad": "Agency staffing summary shows locum nursing coverage added for reopened procedure days.",
                    }
                },
                "search_slack": {
                    "Patient Service Revenue": {
                        "narrow": "Clinic ops confirms expanded procedure slots reopened after provider leave ended.",
                    },
                    "Medical Supplies": {
                        "broad": "Clinical lead notes higher implant and procedure kit usage during the reopened ortho schedule.",
                    }
                },
                "search_calendar": {
                    "Medical Compliance & Audit": {
                        "broad": "Payer documentation audit follow-up scheduled with external compliance team during week four.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Patient Service Revenue",
                    620000,
                    671500,
                    "reopened surgeon schedules increased ortho procedure volume",
                    ["surgeon schedules reopened", "procedure volume above plan", "expanded procedure slots"],
                    ["crm", "slack"],
                    ["crm-patient_service_revenue-broad", "slack-patient_service_revenue-narrow"],
                ),
                supported_line(
                    "Contract Labor — Clinical",
                    134000,
                    156900,
                    "locum nursing coverage was added for the reopened procedure days",
                    ["agency staffing summary", "locum nursing coverage", "reopened procedure days"],
                    ["gmail"],
                    ["gmail-contract_labor_—_clinical-broad"],
                ),
                supported_line(
                    "Medical Supplies",
                    97000,
                    109300,
                    "higher implant and procedure kit usage drove supply spend",
                    ["implant", "procedure kit usage", "reopened ortho schedule"],
                    ["slack"],
                    ["slack-medical_supplies-broad"],
                ),
                supported_line(
                    "Medical Compliance & Audit",
                    18000,
                    24600,
                    "an unplanned payer documentation audit triggered outside compliance review spend",
                    ["unplanned payer documentation audit", "external compliance review"],
                    ["gmail", "calendar"],
                    ["gmail-medical_compliance_&_audit-broad", "calendar-medical_compliance_&_audit-broad"],
                ),
            ],
        },
        {
            "slug": "iona_renewable_storage_november_2024",
            "company": "Iona Renewable Storage",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Professional Fees", "Interconnection Studies", "Travel & Site Visits"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 470000, 448200),
                ("Professional Fees", 21000, 36850),
                ("Interconnection Studies", 54000, 67700),
                ("Travel & Site Visits", 15000, 19200),
                ("Salaries", 186000, 194800),
                ("Software & Subscriptions", 16000, 17150),
                ("Insurance", 14000, 14500),
                ("Permitting Costs", 28000, 32400),
                ("Utilities", 9000, 9350),
                ("Office & Facilities", 12000, 11800),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM notes show one battery project milestone slipped into December after utility review comments extended sign-off.",
                    }
                },
                "search_gmail": {
                    "Professional Fees": {
                        "broad": "Outside counsel and market advisory invoices were approved for a delayed interconnection filing response.",
                    },
                    "Interconnection Studies": {
                        "broad": "Utility consultant change order approved additional modeling iterations for two storage sites.",
                    }
                },
                "search_slack": {
                    "Revenue": {
                        "narrow": "Project channel confirms milestone sign-off moved to early December after utility questions stayed open.",
                    },
                    "Travel & Site Visits": {
                        "broad": "Development lead notes two extra landowner and utility site visits were added during the delay window.",
                    }
                },
                "search_calendar": {
                    "Interconnection Studies": {
                        "broad": "Utility modeling review held twice in late November after revised scenarios were requested.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    470000,
                    448200,
                    "project milestone sign-off slipped into december after utility review delays",
                    ["milestone slipped into december", "utility review comments", "sign-off moved"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "slack-revenue-narrow"],
                    mitigation="track whether December recognizes the delayed milestone",
                ),
                supported_line(
                    "Professional Fees",
                    21000,
                    36850,
                    "outside counsel and market advisory support increased during the delayed filing response",
                    ["outside counsel", "market advisory", "delayed interconnection filing"],
                    ["gmail"],
                    ["gmail-professional_fees-broad"],
                ),
                supported_line(
                    "Interconnection Studies",
                    54000,
                    67700,
                    "additional modeling iterations for two sites drove the change order",
                    ["change order", "additional modeling iterations", "two storage sites"],
                    ["gmail", "calendar"],
                    ["gmail-interconnection_studies-broad", "calendar-interconnection_studies-broad"],
                ),
                supported_line(
                    "Travel & Site Visits",
                    15000,
                    19200,
                    "extra landowner and utility site visits during the delay period increased travel",
                    ["extra landowner", "utility site visits", "delay window"],
                    ["slack"],
                    ["slack-travel_&_site_visits-broad"],
                ),
            ],
        },
        {
            "slug": "juniper_retail_media_november_2024",
            "company": "Juniper Retail Media",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Sales Commissions", "Professional Fees", "Campaign Production"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 430000, 486700),
                ("Sales Commissions", 44000, 54850),
                ("Professional Fees", 17000, 26100),
                ("Campaign Production", 59000, 68400),
                ("Salaries", 188000, 194700),
                ("Ad Serving Fees", 76000, 82100),
                ("Software & Subscriptions", 21000, 21900),
                ("Travel & Entertainment", 7000, 8150),
                ("Insurance", 11000, 10950),
                ("Office & Facilities", 9000, 9200),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM shows two holiday retail campaigns and one budget expansion approved before 30 November, adding +39000 to billings.",
                    }
                },
                "search_gmail": {
                    "Professional Fees": {
                        "broad": "Creative legal review and measurement audit invoices were approved for accelerated holiday campaigns.",
                    }
                },
                "search_slack": {
                    "Sales Commissions": {
                        "broad": "RevOps confirms commission accrual increased with the holiday campaign expansions closed in November.",
                    },
                    "Campaign Production": {
                        "broad": "Production lead says rushed creative refreshes and late client feedback loops required weekend edit sessions.",
                    }
                },
                "search_calendar": {
                    "Campaign Production": {
                        "narrow": "Weekend edit sessions and client approvals were added across the last two November weekends.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    430000,
                    486700,
                    "holiday campaign wins and a budget expansion lifted november billings",
                    ["holiday retail campaigns", "budget expansion", "adding to billings"],
                    ["crm"],
                    ["crm-revenue-broad"],
                ),
                supported_line(
                    "Sales Commissions",
                    44000,
                    54850,
                    "commission accrual rose with november holiday campaign expansions",
                    ["commission accrual", "holiday campaign expansions"],
                    ["slack"],
                    ["slack-sales_commissions-broad"],
                ),
                supported_line(
                    "Professional Fees",
                    17000,
                    26100,
                    "creative legal review and measurement audit support drove the overage",
                    ["creative legal review", "measurement audit", "accelerated holiday campaigns"],
                    ["gmail"],
                    ["gmail-professional_fees-broad"],
                ),
                supported_line(
                    "Campaign Production",
                    59000,
                    68400,
                    "rushed creative refreshes and weekend edit sessions increased production costs",
                    ["rushed creative refreshes", "weekend edit sessions", "late client feedback"],
                    ["slack", "calendar"],
                    ["slack-campaign_production-broad", "calendar-campaign_production-narrow"],
                ),
            ],
        },
        {
            "slug": "keystone_field_operations_november_2024",
            "company": "Keystone Field Operations",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Contract Labor", "Fuel", "Equipment Leases"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 720000, 758400),
                ("Contract Labor", 164000, 188600),
                ("Fuel", 101000, 116900),
                ("Equipment Leases", 58000, 66100),
                ("Salaries", 215000, 224500),
                ("Professional Fees", 19000, 20500),
                ("Travel & Lodging", 26000, 30100),
                ("Software & Subscriptions", 13000, 13250),
                ("Insurance", 14000, 14650),
                ("Repairs & Maintenance", 47000, 50100),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM confirms one storm-response deployment and one utility maintenance expansion billed before month-end.",
                    }
                },
                "search_gmail": {
                    "Contract Labor": {
                        "broad": "Field staffing vendor summary shows incremental crews mobilized for the storm-response deployment.",
                    },
                    "Equipment Leases": {
                        "broad": "Lease amendment invoice covers two additional generators mobilized for the expanded utility scope.",
                    }
                },
                "search_slack": {
                    "Fuel": {
                        "broad": "Ops channel notes higher fleet hours and generator runtime during storm-response work.",
                    },
                    "Revenue": {
                        "narrow": "Commercial lead confirms both storm and utility scope were approved and billed inside November.",
                    }
                },
                "search_calendar": {
                    "Contract Labor": {
                        "broad": "Crew mobilization calls scheduled daily for the first 10 days of the storm-response deployment.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    720000,
                    758400,
                    "storm-response deployment and utility scope expansion lifted november billing",
                    ["storm-response deployment", "utility maintenance expansion", "billed inside november"],
                    ["crm", "slack"],
                    ["crm-revenue-broad", "slack-revenue-narrow"],
                ),
                supported_line(
                    "Contract Labor",
                    164000,
                    188600,
                    "incremental crews mobilized for storm work drove contract labor higher",
                    ["incremental crews", "storm-response deployment", "crew mobilization"],
                    ["gmail", "calendar"],
                    ["gmail-contract_labor-broad", "calendar-contract_labor-broad"],
                ),
                supported_line(
                    "Fuel",
                    101000,
                    116900,
                    "higher fleet hours and generator runtime increased fuel usage",
                    ["higher fleet hours", "generator runtime", "storm-response work"],
                    ["slack"],
                    ["slack-fuel-broad"],
                ),
                supported_line(
                    "Equipment Leases",
                    58000,
                    66100,
                    "two additional generators mobilized under a lease amendment increased lease cost",
                    ["lease amendment invoice", "additional generators", "expanded utility scope"],
                    ["gmail"],
                    ["gmail-equipment_leases-broad"],
                ),
            ],
        },
        {
            "slug": "lumen_education_platform_november_2024",
            "company": "Lumen Education Platform",
            "summary_expectations": {
                "must_surface_lines": ["Revenue", "Customer Support Payroll", "Cloud Hosting", "Print & Swag"],
                "max_opening_drivers": 5,
                "forbid_mixed_total_story": True,
            },
            "accounts": [
                ("Revenue", 510000, 542600),
                ("Customer Support Payroll", 98000, 112900),
                ("Cloud Hosting", 76000, 87300),
                ("Print & Swag", 19000, 26750),
                ("Salaries", 223000, 230400),
                ("Professional Fees", 13000, 14100),
                ("Travel & Entertainment", 8000, 7900),
                ("Marketing Programs", 42000, 44600),
                ("Software & Subscriptions", 16000, 16950),
                ("Insurance", 10000, 9950),
            ],
            "mock_context": {
                "search_crm": {
                    "Revenue": {
                        "broad": "CRM notes show district expansion seats and one multi-campus add-on booked in the last week of November.",
                    }
                },
                "search_gmail": {
                    "Cloud Hosting": {
                        "broad": "Hosting invoice summary shows peak load from expanded seat provisioning and duplicated onboarding environments.",
                    }
                },
                "search_slack": {
                    "Customer Support Payroll": {
                        "broad": "Support lead says weekend onboarding coverage and temporary chat staffing were added for the district expansion wave.",
                    },
                    "Print & Swag": {
                        "broad": "Field marketing thread confirms unplanned launch kits were shipped to all new campus champions.",
                    }
                },
                "search_calendar": {
                    "Customer Support Payroll": {
                        "narrow": "Onboarding command-center shifts were scheduled across two weekends after the district expansion closed.",
                    }
                },
            },
            "oracle_lines": [
                supported_line(
                    "Revenue",
                    510000,
                    542600,
                    "district expansion seats and a multi-campus add-on lifted november revenue",
                    ["district expansion seats", "multi-campus add-on", "last week of november"],
                    ["crm"],
                    ["crm-revenue-broad"],
                ),
                supported_line(
                    "Customer Support Payroll",
                    98000,
                    112900,
                    "weekend onboarding coverage and temporary chat staffing increased payroll",
                    ["weekend onboarding coverage", "temporary chat staffing", "command-center shifts"],
                    ["slack", "calendar"],
                    ["slack-customer_support_payroll-broad", "calendar-customer_support_payroll-narrow"],
                ),
                supported_line(
                    "Cloud Hosting",
                    76000,
                    87300,
                    "expanded seat provisioning and duplicated onboarding environments increased hosting cost",
                    ["expanded seat provisioning", "duplicated onboarding environments", "peak load"],
                    ["gmail"],
                    ["gmail-cloud_hosting-broad"],
                ),
                supported_line(
                    "Print & Swag",
                    19000,
                    26750,
                    "unplanned launch kits for new campus champions caused the overage",
                    ["launch kits", "new campus champions", "shipped"],
                    ["slack"],
                    ["slack-print_&_swag-broad"],
                ),
            ],
        },
    ]


def write_bundle(root: Path, spec: dict) -> None:
    slug = spec["slug"]
    write_workbook(root / f"{slug}.xlsx", spec["company"], spec["accounts"])
    mock_context = {"period": "November 2024", "tool_responses": spec["mock_context"]}
    (root / f"{slug}.mock_context.json").write_text(
        json.dumps(mock_context, indent=2) + "\n",
        encoding="utf-8",
    )
    oracle = {
        "saved_run": "",
        "workbook": f"{slug}.xlsx",
        "period": "November 2024",
        "summary_expectations": spec["summary_expectations"],
        "lines": spec["oracle_lines"],
    }
    (root / f"{slug}.oracle.json").write_text(
        json.dumps(oracle, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    root = Path(__file__).resolve().parent
    root.mkdir(parents=True, exist_ok=True)
    specs = build_specs()[:10]
    for spec in specs:
        write_bundle(root, spec)
    manifest = {
        "period": "November 2024",
        "bundle_count": len(specs),
        "bundles": [
            {
                "slug": spec["slug"],
                "xlsx": f"{spec['slug']}.xlsx",
                "mock_context": f"{spec['slug']}.mock_context.json",
                "oracle": f"{spec['slug']}.oracle.json",
            }
            for spec in specs
        ],
    }
    (root / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
